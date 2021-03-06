# -*- coding: utf-8 -*-

from django.conf import settings
from django.contrib import admin
from django.utils.text import slugify
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl import Workbook

from erudit.models import Journal

from .models import JournalRoyalty
from .models import RoyaltyReport


class RoyaltyReportAdmin(admin.ModelAdmin):
    def save_model(self, request, obj, form, change):
        obj.save()
        self.create_journal_royalties(obj)

    def create_journal_royalties(self, report):
        wb = load_workbook(report.report_file.path)

        # First step: removes od journal royalties associated with the RoyaltyReport instance.
        report.journalroyalty_set.all().delete()

        for sheet_name in wb.get_sheet_names():
            ws = wb.get_sheet_by_name(sheet_name)
            journal_code_or_lid = ws['H2'].value
            journal = Journal.objects \
                .filter(Q(code=journal_code_or_lid) | Q(localidentifier=journal_code_or_lid)) \
                .first()

            if not journal_code_or_lid or journal is None:
                continue

            # Constructs a new workbook by removing all the sheets that are not associated with the
            # current sheet.
            journal_wb = Workbook()
            journal_ws = journal_wb.create_sheet(journal_code_or_lid, 0)
            for row in ws.rows:
                for cell in row:
                    journal_ws[cell.coordinate] = cell.value
                    journal_ws[cell.coordinate].data_type = cell.data_type
                    if cell.has_style:
                        journal_ws[cell.coordinate].alignment = cell.alignment.copy()
                        journal_ws[cell.coordinate].font = cell.font.copy()
                        journal_ws[cell.coordinate].number_format = cell.number_format
            journal_ws.column_dimensions = ws.column_dimensions
            journal_ws.row_dimensions = ws.row_dimensions

            workbook_file_name = 'royalty_reports/' \
                + 'Redevances_{nameslug}_{enddate}.xlsx'.format(
                    nameslug=slugify(journal.name)[:30], enddate=report.end)
            journal_wb.save(settings.MEDIA_ROOT + '/' + workbook_file_name)

            journal_royalty = JournalRoyalty()
            journal_royalty.royalty_report = report
            journal_royalty.journal = journal
            journal_royalty.report_file = workbook_file_name
            journal_royalty.save()


class JournalRoyaltyAdmin(admin.ModelAdmin):
    list_display = ['royalty_report', 'journal', ]


admin.site.register(RoyaltyReport, RoyaltyReportAdmin)
admin.site.register(JournalRoyalty, JournalRoyaltyAdmin)
